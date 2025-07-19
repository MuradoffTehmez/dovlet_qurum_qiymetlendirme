# core/reports.py
"""
Q360 Hesabat Mərkəzi
Dynamic filtrlər, multiple formatlar və smart caching ilə hesabat sistemi
"""

from django.db.models import Count, Avg, Q, Max, Min
from django.utils import timezone
from django.core.cache import cache
from django.template.loader import render_to_string
from django.http import HttpResponse
from datetime import datetime, timedelta
import json
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib import colors
from reportlab.lib.units import inch

from .models import (
    Ishchi, Qiymetlendirme, InkishafPlani, Hedef, 
    OrganizationUnit, QiymetlendirmeDovru, Notification
)

class ReportManager:
    """Hesabat idarəetmə mərkəzi"""
    
    @staticmethod
    def get_available_reports():
        """Mövcud hesabat növlərini qaytarır"""
        return {
            'employee_performance': {
                'name': 'İşçi Performans Hesabatı',
                'description': 'İşçilərin performans qiymətləndirmələri və statistikaları',
                'icon': 'fas fa-user-chart',
                'formats': ['pdf', 'excel', 'csv']
            },
            'department_summary': {
                'name': 'Şöbə Xülasə Hesabatı', 
                'description': 'Şöbələr üzrə ümumi performans və məlumatlar',
                'icon': 'fas fa-building',
                'formats': ['pdf', 'excel', 'csv']
            },
            'development_plans': {
                'name': 'İnkişaf Planları Hesabatı',
                'description': 'Aktiv inkişaf planları və hədəflərin statusu',
                'icon': 'fas fa-target',
                'formats': ['pdf', 'excel']
            },
            'evaluation_period': {
                'name': 'Qiymətləndirmə Dövrü Hesabatı',
                'description': 'Seçilmiş dövr üzrə tam qiymətləndirmə hesabatı',
                'icon': 'fas fa-calendar-check',
                'formats': ['pdf', 'excel']
            },
            'notification_analytics': {
                'name': 'Bildiriş Analitikası',
                'description': 'Bildiriş sistemi statistikaları və trendlər',
                'icon': 'fas fa-bell',
                'formats': ['pdf', 'excel', 'csv']
            }
        }
    
    @staticmethod
    def get_filter_options():
        """Filtr seçimlərini qaytarır"""
        return {
            'date_ranges': [
                ('last_7_days', 'Son 7 gün'),
                ('last_30_days', 'Son 30 gün'),
                ('last_3_months', 'Son 3 ay'),
                ('last_6_months', 'Son 6 ay'),
                ('last_year', 'Son il'),
                ('custom', 'Xüsusi tarix aralığı')
            ],
            'departments': list(OrganizationUnit.objects.values_list('id', 'name')),
            'evaluation_periods': list(QiymetlendirmeDovru.objects.values_list('id', 'ad')),
            'roles': list(Ishchi.Rol.choices),
            'performance_ranges': [
                ('excellent', '90-100 (Əla)'),
                ('good', '70-89 (Yaxşı)'),
                ('average', '50-69 (Orta)'),
                ('poor', '0-49 (Zəif)')
            ]
        }


class EmployeePerformanceReport:
    """İşçi Performans Hesabatı"""
    
    def __init__(self, filters=None):
        self.filters = filters or {}
        
    def get_data(self):
        """Hesabat məlumatlarını hazırlayır"""
        cache_key = f"employee_performance_report_{hash(str(self.filters))}"
        data = cache.get(cache_key)
        
        if data is None:
            data = self._generate_data()
            cache.set(cache_key, data, 300)  # 5 dəqiqə cache
            
        return data
    
    def _generate_data(self):
        """Məlumatları generasiya edir"""
        # Base query
        evaluations = Qiymetlendirme.objects.select_related(
            'qiymetlendirilen', 'qiymetlendiren', 'dovr'
        )
        
        # Filtrlər tətbiq et
        evaluations = self._apply_filters(evaluations)
        
        # Statistikalar hesabla
        stats = self._calculate_statistics(evaluations)
        
        # Detallı məlumatlar
        detailed_data = self._prepare_detailed_data(evaluations)
        
        return {
            'stats': stats,
            'detailed_data': detailed_data,
            'filters_applied': self.filters,
            'generated_at': timezone.now(),
            'total_evaluations': evaluations.count()
        }
    
    def _apply_filters(self, queryset):
        """Filtrləri tətbiq edir"""
        if 'date_from' in self.filters and self.filters['date_from']:
            queryset = queryset.filter(tarix__gte=self.filters['date_from'])
            
        if 'date_to' in self.filters and self.filters['date_to']:
            queryset = queryset.filter(tarix__lte=self.filters['date_to'])
            
        if 'department' in self.filters and self.filters['department']:
            queryset = queryset.filter(
                qiymetlendirilen__organization_unit_id=self.filters['department']
            )
            
        if 'role' in self.filters and self.filters['role']:
            queryset = queryset.filter(qiymetlendirilen__rol=self.filters['role'])
            
        if 'evaluation_period' in self.filters and self.filters['evaluation_period']:
            queryset = queryset.filter(dovr_id=self.filters['evaluation_period'])
            
        return queryset
    
    def _calculate_statistics(self, evaluations):
        """Statistikaları hesablayır"""
        if not evaluations.exists():
            return {
                'total_employees': 0,
                'average_score': 0,
                'performance_distribution': {},
                'top_performers': [],
                'departments_comparison': {}
            }
        
        # Ümumi statistikalar
        avg_score = evaluations.aggregate(Avg('umumi_qiymet'))['umumi_qiymet__avg'] or 0
        
        # Performans paylanması
        performance_dist = {
            'excellent': evaluations.filter(umumi_qiymet__gte=90).count(),
            'good': evaluations.filter(umumi_qiymet__gte=70, umumi_qiymet__lt=90).count(),
            'average': evaluations.filter(umumi_qiymet__gte=50, umumi_qiymet__lt=70).count(),
            'poor': evaluations.filter(umumi_qiymet__lt=50).count(),
        }
        
        # Top performerlər
        top_performers = evaluations.filter(
            umumi_qiymet__gte=85
        ).order_by('-umumi_qiymet')[:10]
        
        # Şöbələr müqayisəsi
        dept_comparison = evaluations.values(
            'qiymetlendirilen__organization_unit__name'
        ).annotate(
            avg_score=Avg('umumi_qiymet'),
            employee_count=Count('qiymetlendirilen', distinct=True)
        ).order_by('-avg_score')
        
        return {
            'total_employees': evaluations.values('qiymetlendirilen').distinct().count(),
            'average_score': round(avg_score, 2),
            'performance_distribution': performance_dist,
            'top_performers': list(top_performers),
            'departments_comparison': list(dept_comparison)
        }
    
    def _prepare_detailed_data(self, evaluations):
        """Detallı məlumatları hazırlayır"""
        return evaluations.values(
            'qiymetlendirilen__first_name',
            'qiymetlendirilen__last_name',
            'qiymetlendirilen__organization_unit__name',
            'qiymetlendirilen__rol',
            'umumi_qiymet',
            'tarix',
            'qiymetlendiren__first_name',
            'qiymetlendiren__last_name'
        ).order_by('-umumi_qiymet')


class ReportGenerator:
    """Hesabat generasiya sinifi"""
    
    @staticmethod
    def generate_pdf(report_data, report_type):
        """PDF hesabat generasiya edir"""
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []
        
        # Başlıq
        title = Paragraph(f"Q360 - {report_data.get('title', 'Hesabat')}", styles['Title'])
        story.append(title)
        story.append(Spacer(1, 12))
        
        # Generasiya tarixi
        gen_date = report_data['generated_at'].strftime('%d.%m.%Y %H:%M')
        date_para = Paragraph(f"Generasiya tarixi: {gen_date}", styles['Normal'])
        story.append(date_para)
        story.append(Spacer(1, 12))
        
        if report_type == 'employee_performance':
            story.extend(ReportGenerator._create_performance_pdf_content(report_data, styles))
        elif report_type == 'department_summary':
            story.extend(ReportGenerator._create_department_pdf_content(report_data, styles))
        
        doc.build(story)
        buffer.seek(0)
        return buffer
    
    @staticmethod
    def _create_performance_pdf_content(data, styles):
        """Performans hesabatı üçün PDF məzmunu"""
        content = []
        stats = data['stats']
        
        # Ümumi statistikalar
        content.append(Paragraph("Ümumi Statistikalar", styles['Heading2']))
        
        stats_table_data = [
            ['Metrика', 'Dəyər'],
            ['Cəmi İşçi Sayı', str(stats['total_employees'])],
            ['Orta Performans Balı', f"{stats['average_score']:.2f}"],
            ['Cəmi Qiymətləndirmə', str(data['total_evaluations'])]
        ]
        
        stats_table = Table(stats_table_data, colWidths=[3*inch, 2*inch])
        stats_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        content.append(stats_table)
        content.append(Spacer(1, 12))
        
        # Performans paylanması
        content.append(Paragraph("Performans Paylanması", styles['Heading2']))
        
        perf_data = [
            ['Səviyyə', 'Sayı', 'Faiz'],
            ['Əla (90-100)', str(stats['performance_distribution']['excellent']), 
             f"{(stats['performance_distribution']['excellent']/data['total_evaluations']*100):.1f}%" if data['total_evaluations'] > 0 else "0%"],
            ['Yaxşı (70-89)', str(stats['performance_distribution']['good']),
             f"{(stats['performance_distribution']['good']/data['total_evaluations']*100):.1f}%" if data['total_evaluations'] > 0 else "0%"],
            ['Orta (50-69)', str(stats['performance_distribution']['average']),
             f"{(stats['performance_distribution']['average']/data['total_evaluations']*100):.1f}%" if data['total_evaluations'] > 0 else "0%"],
            ['Zəif (0-49)', str(stats['performance_distribution']['poor']),
             f"{(stats['performance_distribution']['poor']/data['total_evaluations']*100):.1f}%" if data['total_evaluations'] > 0 else "0%"]
        ]
        
        perf_table = Table(perf_data, colWidths=[2*inch, 1*inch, 1*inch])
        perf_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        content.append(perf_table)
        
        return content
    
    @staticmethod
    def generate_excel(report_data, report_type):
        """Excel hesabat generasiya edir"""
        buffer = BytesIO()
        workbook = openpyxl.Workbook()
        
        if report_type == 'employee_performance':
            ReportGenerator._create_performance_excel(workbook, report_data)
        elif report_type == 'department_summary':
            ReportGenerator._create_department_excel(workbook, report_data)
        
        workbook.save(buffer)
        buffer.seek(0)
        return buffer
    
    @staticmethod
    def _create_performance_excel(workbook, data):
        """Performans hesabatı üçün Excel yaradır"""
        ws = workbook.active
        ws.title = "Performans Hesabatı"
        
        # Başlıq stilləri
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Başlıq
        ws['A1'] = "Q360 Performans Hesabatı"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:E1')
        
        # Generasiya tarixi
        ws['A2'] = f"Generasiya tarixi: {data['generated_at'].strftime('%d.%m.%Y %H:%M')}"
        ws.merge_cells('A2:E2')
        
        # Statistikalar bölməsi
        ws['A4'] = "Ümumi Statistikalar"
        ws['A4'].font = Font(bold=True, size=14)
        
        stats = data['stats']
        stats_data = [
            ['Metrика', 'Dəyər'],
            ['Cəmi İşçi Sayı', stats['total_employees']],
            ['Orta Performans Balı', round(stats['average_score'], 2)],
            ['Cəmi Qiymətləndirmə', data['total_evaluations']]
        ]
        
        for row_idx, row_data in enumerate(stats_data, start=5):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 5:  # Header row
                    cell.font = header_font
                    cell.fill = header_fill
        
        # Detallı məlumatlar
        if data['detailed_data']:
            ws['A10'] = "Detallı Məlumatlar"
            ws['A10'].font = Font(bold=True, size=14)
            
            headers = ['Ad', 'Soyad', 'Şöbə', 'Rol', 'Performans Balı', 'Tarix', 'Qiymətləndirən']
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=11, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
            
            for row_idx, item in enumerate(data['detailed_data'], start=12):
                ws.cell(row=row_idx, column=1, value=item['qiymetlendirilen__first_name'])
                ws.cell(row=row_idx, column=2, value=item['qiymetlendirilen__last_name'])
                ws.cell(row=row_idx, column=3, value=item['qiymetlendirilen__organization_unit__name'])
                ws.cell(row=row_idx, column=4, value=item['qiymetlendirilen__rol'])
                ws.cell(row=row_idx, column=5, value=item['umumi_qiymet'])
                ws.cell(row=row_idx, column=6, value=item['tarix'].strftime('%d.%m.%Y') if item['tarix'] else '')
                ws.cell(row=row_idx, column=7, value=f"{item['qiymetlendiren__first_name']} {item['qiymetlendiren__last_name']}")
        
        # Sütun eni təyin et
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    @staticmethod
    def generate_csv(report_data, report_type):
        """CSV hesabat generasiya edir"""
        import csv
        from io import StringIO
        
        output = StringIO()
        writer = csv.writer(output)
        
        if report_type == 'employee_performance':
            # CSV başlıqları
            writer.writerow(['Ad', 'Soyad', 'Şöbə', 'Rol', 'Performans Balı', 'Tarix', 'Qiymətləndirən'])
            
            # Məlumatlar
            for item in report_data['detailed_data']:
                writer.writerow([
                    item['qiymetlendirilen__first_name'],
                    item['qiymetlendirilen__last_name'],
                    item['qiymetlendirilen__organization_unit__name'],
                    item['qiymetlendirilen__rol'],
                    item['umumi_qiymet'],
                    item['tarix'].strftime('%d.%m.%Y') if item['tarix'] else '',
                    f"{item['qiymetlendiren__first_name']} {item['qiymetlendiren__last_name']}"
                ])
        
        output.seek(0)
        return output


class ReportScheduler:
    """Planlanmış hesabatlar"""
    
    @staticmethod
    def schedule_weekly_report():
        """Həftəlik avtomatik hesabat"""
        from .tasks import generate_report_task
        from .notifications import NotificationManager
        
        # HR Manager-lara həftəlik hesabat göndər
        hr_managers = Ishchi.objects.filter(groups__name='HR Manager')
        
        for manager in hr_managers:
            generate_report_task.delay(
                report_type='employee_performance',
                user_id=manager.id,
                filters={'date_range': 'last_7_days'}
            )
            
            NotificationManager.create_and_send(
                recipient=manager,
                title="Həftəlik Performans Hesabatı Hazır",
                message="Yeni həftəlik performans hesabatı generasiya edildi və e-poçtunuza göndərildi.",
                notification_type='INFO',
                priority='MEDIUM',
                send_email=True
            )