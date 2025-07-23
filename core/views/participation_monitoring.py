"""
Participation Monitoring (İştirak Nəzarəti) Views
Qiymətləndirmə prosesində iştirak faizini izləyir və xatırlatmalar göndərir
"""
from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.db.models import Count, Q, F, Max
from django.db import models
from django.http import JsonResponse
from django.contrib import messages
from django.utils import timezone
from datetime import timedelta
import json

from ..models import (
    Ishchi, Qiymetlendirme, QiymetlendirmeDovru, 
    OrganizationUnit, Notification
)
from ..permissions import require_role
from ..tasks import send_notification_email_task


@login_required
@require_role(['ADMIN', 'SUPERADMIN', 'REHBER'])
def participation_dashboard(request):
    """İştirak nəzarəti ana səhifəsi"""
    
    # Aktiv dövrləri tap
    active_cycles = QiymetlendirmeDovru.objects.filter(
        aktivdir=True,
        bashlama_tarixi__lte=timezone.now().date(),
        bitme_tarixi__gte=timezone.now().date()
    ).order_by('-bashlama_tarixi')
    
    # Seçilmiş dövr
    cycle_id = request.GET.get('cycle_id')
    if cycle_id:
        selected_cycle = get_object_or_404(QiymetlendirmeDovru, id=cycle_id)
    elif active_cycles.exists():
        selected_cycle = active_cycles.first()
    else:
        selected_cycle = None
    
    participation_data = None
    if selected_cycle:
        participation_data = get_participation_statistics(selected_cycle)
    
    context = {
        'active_cycles': active_cycles,
        'selected_cycle': selected_cycle,
        'participation_data': participation_data,
        'page_title': 'İştirak Nəzarəti'
    }
    
    return render(request, 'participation_monitoring/dashboard.html', context)


@login_required
@require_role(['ADMIN', 'SUPERADMIN', 'REHBER'])
def participation_details(request, cycle_id):
    """Müəyyən dövr üçün detallı iştirak məlumatları"""
    
    cycle = get_object_or_404(QiymetlendirmeDovru, id=cycle_id)
    department_id = request.GET.get('department_id')
    
    # Şöbə filtri
    employees_query = Ishchi.objects.filter(is_active=True)
    if department_id:
        employees_query = employees_query.filter(organization_unit_id=department_id)
    
    # Hər işçi üçün iştirak statusu - Optimized single query approach
    participation_details = []
    
    # Single query to get all employee statistics at once
    employee_stats = employees_query.annotate(
        assigned_count=Count(
            'etdiyi_qiymetler',
            filter=Q(etdiyi_qiymetler__dovr=cycle),
            distinct=True
        ),
        completed_count=Count(
            'etdiyi_qiymetler',
            filter=Q(
                etdiyi_qiymetler__dovr=cycle,
                etdiyi_qiymetler__status=Qiymetlendirme.Status.TAMAMLANDI
            ),
            distinct=True
        )
    ).prefetch_related('etdiyi_qiymetler')
    
    # Get last activity dates in a single query
    last_activities = {}
    last_activity_data = Qiymetlendirme.objects.filter(
        qiymetlendiren__in=employees_query,
        dovr=cycle,
        status=Qiymetlendirme.Status.TAMAMLANDI
    ).values('qiymetlendiren_id').annotate(
        last_activity=Max('tamamlanma_tarixi')
    )
    
    for activity in last_activity_data:
        last_activities[activity['qiymetlendiren_id']] = activity['last_activity']
    
    for employee in employee_stats:
        assigned_count = employee.assigned_count
        completed_count = employee.completed_count
        pending_count = assigned_count - completed_count
        
        # İştirak faizi
        if assigned_count > 0:
            participation_rate = (completed_count / assigned_count) * 100
        else:
            participation_rate = 0
        
        # Last activity date
        last_activity = last_activities.get(employee.id)
        if last_activity:
            last_activity = last_activity.date()
        
        participation_details.append({
            'employee': employee,
            'assigned_count': assigned_count,
            'completed_count': completed_count,
            'pending_count': pending_count,
            'participation_rate': round(participation_rate, 1),
            'status': get_participation_status(participation_rate),
            'last_activity': last_activity
        })
    
    # Sıralama
    sort_by = request.GET.get('sort', 'participation_rate')
    if sort_by == 'participation_rate':
        participation_details.sort(key=lambda x: x['participation_rate'])
    elif sort_by == 'pending_count':
        participation_details.sort(key=lambda x: x['pending_count'], reverse=True)
    elif sort_by == 'last_activity':
        participation_details.sort(key=lambda x: x['last_activity'] or timezone.now().date(), reverse=True)
    
    # Şöbələr siyahısı
    departments = OrganizationUnit.objects.filter(
        ishchiler__isnull=False
    ).distinct().order_by('name')
    
    context = {
        'cycle': cycle,
        'participation_details': participation_details,
        'departments': departments,
        'selected_department_id': int(department_id) if department_id else None,
        'sort_by': sort_by,
        'page_title': f'İştirak Detalları - {cycle.ad}'
    }
    
    return render(request, 'participation_monitoring/details.html', context)


@login_required
@require_role(['ADMIN', 'SUPERADMIN', 'REHBER'])
def send_participation_reminders(request, cycle_id):
    """İştirak xatırlatmaları göndər"""
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Yalnız POST metodu'})
    
    cycle = get_object_or_404(QiymetlendirmeDovru, id=cycle_id)
    reminder_type = request.POST.get('reminder_type', 'all')  # all, low_participation, no_participation
    
    # Hədəf işçiləri müəyyən et
    if reminder_type == 'no_participation':
        # Heç bir qiymətləndirmə tamamlamamış işçilər
        target_employees = get_non_participating_employees(cycle)
    elif reminder_type == 'low_participation':
        # Aşağı iştirak faizli işçilər (50%-dən az)
        target_employees = get_low_participation_employees(cycle)
    else:
        # Bütün natamam işçilər
        target_employees = get_incomplete_participation_employees(cycle)
    
    # Xatırlatma göndər
    sent_count = 0
    for employee in target_employees:
        try:
            # Bildiriş yarat
            Notification.create_notification(
                recipient=employee,
                title=f"Qiymətləndirmə Xatırlatması - {cycle.ad}",
                message=f"Hörmətli {employee.get_full_name()}, {cycle.ad} dövrü üçün qiymətləndirmələrinizi tamamlamağınız xahiş olunur. Son tarix: {cycle.bitme_tarixi.strftime('%d.%m.%Y')}",
                notification_type=Notification.NotificationType.DEADLINE_REMINDER,
                priority=Notification.Priority.HIGH,
                action_url="/dashboard/",
                action_text="Qiymətləndirmələrə Bax"
            )
            
            # E-mail göndər (asinxron)
            send_notification_email_task.delay(
                employee.email,
                'deadline_reminder',
                {
                    'user_name': employee.get_full_name(),
                    'cycle_name': cycle.ad,
                    'deadline': cycle.bitme_tarixi.strftime('%d.%m.%Y'),
                    'subject': f'Qiymətləndirmə Xatırlatması - {cycle.ad}'
                }
            )
            
            sent_count += 1
            
        except:
            continue
    
    messages.success(request, f"{sent_count} işçiyə xatırlatma göndərildi.")
    
    return JsonResponse({
        'success': True,
        'sent_count': sent_count,
        'message': f"{sent_count} işçiyə xatırlatma göndərildi."
    })


@login_required
@require_role(['ADMIN', 'SUPERADMIN'])
def participation_analytics_api(request, cycle_id):
    """İştirak analitikası üçün API"""
    
    cycle = get_object_or_404(QiymetlendirmeDovru, id=cycle_id)
    
    # Əsas statistikalar
    stats = get_participation_statistics(cycle)
    
    # Şöbələr üzrə breakdown
    departments_data = []
    departments = OrganizationUnit.objects.filter(
        ishchiler__isnull=False
    ).distinct()
    
    for dept in departments:
        dept_stats = get_department_participation_stats(dept, cycle)
        if dept_stats['total_employees'] > 0:
            departments_data.append(dept_stats)
    
    # Chart üçün məlumatlar
    api_data = {
        'overall_stats': stats,
        'departments': departments_data,
        'timeline_data': get_participation_timeline(cycle),
        'reminder_history': get_reminder_history(cycle)
    }
    
    return JsonResponse(api_data)


@login_required
@require_role(['ADMIN', 'SUPERADMIN'])
def participation_export(request, cycle_id):
    """İştirak məlumatlarını Excel formatında ixrac et"""
    
    cycle = get_object_or_404(QiymetlendirmeDovru, id=cycle_id)
    
    # Excel yaratma
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from django.http import HttpResponse
    from io import BytesIO
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "İştirak Hesabatı"
    
    # Başlıq
    ws['A1'] = f"İştirak Hesabatı - {cycle.ad}"
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:G1')
    
    # Sütun başlıqları
    headers = ['Ad Soyad', 'Şöbə', 'Təyin Edilmiş', 'Tamamlanmış', 'Qalan', 'İştirak Faizi', 'Status']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Məlumatlar
    employees = Ishchi.objects.filter(is_active=True).order_by('first_name', 'last_name')
    
    for row, employee in enumerate(employees, 4):
        assigned_evaluations = Qiymetlendirme.objects.filter(
            qiymetlendiren=employee,
            dovr=cycle
        )
        
        completed_evaluations = assigned_evaluations.filter(
            status=Qiymetlendirme.Status.TAMAMLANDI
        )
        
        assigned_count = assigned_evaluations.count()
        completed_count = completed_evaluations.count()
        pending_count = assigned_count - completed_count
        
        if assigned_count > 0:
            participation_rate = (completed_count / assigned_count) * 100
        else:
            participation_rate = 0
        
        status = get_participation_status(participation_rate)
        
        ws.cell(row=row, column=1, value=employee.get_full_name())
        ws.cell(row=row, column=2, value=employee.organization_unit.name if employee.organization_unit else "")
        ws.cell(row=row, column=3, value=assigned_count)
        ws.cell(row=row, column=4, value=completed_count)
        ws.cell(row=row, column=5, value=pending_count)
        ws.cell(row=row, column=6, value=f"{participation_rate:.1f}%")
        ws.cell(row=row, column=7, value=status)
    
    # Sütun eni
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Response
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response = HttpResponse(
        buffer.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="istirak_hesabati_{cycle.ad}.xlsx"'
    
    return response


def get_participation_statistics(cycle):
    """Dövr üçün ümumi iştirak statistikalarını hesablayır"""
    
    # Aktiv işçilər
    total_employees = Ishchi.objects.filter(is_active=True).count()
    
    # Bu dövrda ən azı bir qiymətləndirmə tapşırığı olan işçilər
    evaluators = Ishchi.objects.filter(
        etdiyi_qiymetler__dovr=cycle
    ).distinct()
    
    # Ən azı bir qiymətləndirmə tamamlamış işçilər
    active_evaluators = evaluators.filter(
        etdiyi_qiymetler__dovr=cycle,
        etdiyi_qiymetler__status=Qiymetlendirme.Status.TAMAMLANDI
    ).distinct()
    
    # Bütün qiymətləndirmələrini tamamlamış işçilər
    fully_completed = []
    for evaluator in evaluators:
        assigned = Qiymetlendirme.objects.filter(qiymetlendiren=evaluator, dovr=cycle)
        completed = assigned.filter(status=Qiymetlendirme.Status.TAMAMLANDI)
        
        if assigned.count() > 0 and assigned.count() == completed.count():
            fully_completed.append(evaluator)
    
    # Faizlər
    total_evaluators = evaluators.count()
    participation_rate = (active_evaluators.count() / total_evaluators * 100) if total_evaluators > 0 else 0
    completion_rate = (len(fully_completed) / total_evaluators * 100) if total_evaluators > 0 else 0
    
    return {
        'total_employees': total_employees,
        'total_evaluators': total_evaluators,
        'active_evaluators': active_evaluators.count(),
        'fully_completed': len(fully_completed),
        'participation_rate': round(participation_rate, 1),
        'completion_rate': round(completion_rate, 1),
        'days_remaining': (cycle.bitme_tarixi - timezone.now().date()).days,
        'cycle_progress': calculate_cycle_progress(cycle)
    }


def get_department_participation_stats(department, cycle):
    """Şöbə üçün iştirak statistikalarını hesablayır"""
    
    dept_employees = Ishchi.objects.filter(
        organization_unit=department,
        is_active=True
    )
    
    dept_evaluators = dept_employees.filter(
        etdiyi_qiymetler__dovr=cycle
    ).distinct()
    
    # Aktiv iştirak edənlər
    active_dept_evaluators = dept_evaluators.filter(
        etdiyi_qiymetler__dovr=cycle,
        etdiyi_qiymetler__status=Qiymetlendirme.Status.TAMAMLANDI
    ).distinct()
    
    # Faiz hesabla
    total_evaluators = dept_evaluators.count()
    participation_rate = (active_dept_evaluators.count() / total_evaluators * 100) if total_evaluators > 0 else 0
    
    return {
        'department_name': department.name,
        'total_employees': dept_employees.count(),
        'total_evaluators': total_evaluators,
        'active_evaluators': active_dept_evaluators.count(),
        'participation_rate': round(participation_rate, 1)
    }


def get_non_participating_employees(cycle):
    """Heç bir qiymətləndirmə tamamlamamış işçiləri qaytarır"""
    
    return Ishchi.objects.filter(
        etdiyi_qiymetler__dovr=cycle,
        is_active=True
    ).exclude(
        etdiyi_qiymetler__dovr=cycle,
        etdiyi_qiymetler__status=Qiymetlendirme.Status.TAMAMLANDI
    ).distinct()


def get_low_participation_employees(cycle, threshold=50):
    """Aşağı iştirak faizli işçiləri qaytarır"""
    
    low_participation = []
    
    evaluators = Ishchi.objects.filter(
        etdiyi_qiymetler__dovr=cycle,
        is_active=True
    ).distinct()
    
    for evaluator in evaluators:
        assigned = Qiymetlendirme.objects.filter(qiymetlendiren=evaluator, dovr=cycle)
        completed = assigned.filter(status=Qiymetlendirme.Status.TAMAMLANDI)
        
        if assigned.count() > 0:
            participation_rate = (completed.count() / assigned.count()) * 100
            if participation_rate < threshold:
                low_participation.append(evaluator)
    
    return low_participation


def get_incomplete_participation_employees(cycle):
    """Natamam qiymətləndirməsi olan işçiləri qaytarır"""
    
    incomplete = []
    
    evaluators = Ishchi.objects.filter(
        etdiyi_qiymetler__dovr=cycle,
        is_active=True
    ).distinct()
    
    for evaluator in evaluators:
        assigned = Qiymetlendirme.objects.filter(qiymetlendiren=evaluator, dovr=cycle)
        completed = assigned.filter(status=Qiymetlendirme.Status.TAMAMLANDI)
        
        if assigned.count() > completed.count():
            incomplete.append(evaluator)
    
    return incomplete


def get_participation_status(participation_rate):
    """İştirak faizinə əsasən status qaytarır"""
    
    if participation_rate >= 100:
        return "Tamamlandı"
    elif participation_rate >= 75:
        return "Yaxşı"
    elif participation_rate >= 50:
        return "Orta"
    elif participation_rate > 0:
        return "Aşağı"
    else:
        return "Başlanmayıb"


def get_last_activity_date(employee, cycle):
    """İşçinin son qiymətləndirmə aktivlik tarixini qaytarır"""
    
    last_evaluation = Qiymetlendirme.objects.filter(
        qiymetlendiren=employee,
        dovr=cycle,
        status=Qiymetlendirme.Status.TAMAMLANDI
    ).order_by('-tamamlanma_tarixi').first()
    
    if last_evaluation and last_evaluation.tamamlanma_tarixi:
        return last_evaluation.tamamlanma_tarixi.date()
    
    return None


def get_participation_timeline(cycle):
    """Dövr ərzində iştirak timeline-ını qaytarır"""
    
    # Hər gün üçün tamamlanmış qiymətləndirmələr sayı
    from django.db.models import DateField
    from django.db.models.functions import Cast
    
    timeline_data = []
    current_date = cycle.bashlama_tarixi
    
    while current_date <= min(cycle.bitme_tarixi, timezone.now().date()):
        completed_that_day = Qiymetlendirme.objects.filter(
            dovr=cycle,
            status=Qiymetlendirme.Status.TAMAMLANDI,
            tamamlanma_tarixi__date=current_date
        ).count()
        
        timeline_data.append({
            'date': current_date.strftime('%Y-%m-%d'),
            'completed_count': completed_that_day
        })
        
        current_date += timedelta(days=1)
    
    return timeline_data


def get_reminder_history(cycle):
    """Xatırlatma tarixçəsini qaytarır"""
    
    # Deadline reminder bildirişləri
    reminders = Notification.objects.filter(
        notification_type=Notification.NotificationType.DEADLINE_REMINDER,
        created_at__gte=cycle.bashlama_tarixi,
        created_at__lte=timezone.now(),
        message__icontains=cycle.ad
    ).values('created_at__date').annotate(
        count=Count('id')
    ).order_by('created_at__date')
    
    return [{
        'date': item['created_at__date'].strftime('%Y-%m-%d'),
        'reminder_count': item['count']
    } for item in reminders]


def calculate_cycle_progress(cycle):
    """Dövrün tamamlanma faizini hesablayır"""
    
    total_days = (cycle.bitme_tarixi - cycle.bashlama_tarixi).days
    elapsed_days = (timezone.now().date() - cycle.bashlama_tarixi).days
    
    if total_days <= 0:
        return 100
    
    progress = (elapsed_days / total_days) * 100
    return min(max(progress, 0), 100)


def main_function(*args, **kwargs):
    """Placeholder for main participation monitoring function (not implemented)."""
    # TODO: Implement actual logic
    raise NotImplementedError("main_function in participation_monitoring is not implemented yet")